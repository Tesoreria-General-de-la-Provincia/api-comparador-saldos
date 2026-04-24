"""
Módulo para generar archivos Excel desde DataFrames.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import numpy as np


def dataframe_to_excel_bytes(
    df: pd.DataFrame, sheet_name: str = "Sheet1", currency_cols: set[str] | None = None
) -> bytes:
    """
    Convierte un DataFrame a bytes de archivo Excel (.xlsx).

    El archivo generado incluye:
    - Headers en negrita
    - Valores NaN representados como celdas vacías
    - Números formateados correctamente
    - Ajuste automático de ancho de columnas

    Args:
        df: DataFrame a convertir
        sheet_name: Nombre de la hoja de Excel

    Returns:
        bytes: Contenido del archivo Excel en formato bytes

    Example:
        >>> df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
        >>> excel_bytes = dataframe_to_excel_bytes(df, 'MiHoja')
        >>> with open('output.xlsx', 'wb') as f:
        ...     f.write(excel_bytes)
    """
    # Crear un nuevo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Estilos
    header_fill = PatternFill(
        start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"
    )
    header_font = Font(bold=True)
    zebra_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    safyc_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    border_side = Side(style="thin", color="000000")
    cell_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )

    # Escribir headers (primera fila)
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = cell_border

    # Escribir datos (desde fila 2)
    safyc_col_idx = None
    if "COD_CTA_SAFYC" in headers:
        safyc_col_idx = headers.index("COD_CTA_SAFYC") + 1

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        apply_zebra = row_idx % 2 == 0
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Manejar valores NaN y None
            if pd.isna(value) or value is None:
                cell.value = None  # Celda vacía en Excel
            # Manejar valores numéricos
            elif isinstance(value, (int, float, np.integer, np.floating)):
                # Convertir numpy types a Python types
                if isinstance(value, (np.integer, np.int64, np.int32)):
                    cell.value = int(value)
                elif isinstance(value, (np.floating, np.float64, np.float32)):
                    if np.isnan(value):
                        cell.value = None
                    else:
                        # mantener el valor numérico tal cual para que Excel lo formatee
                        cell.value = float(value)
                else:
                    cell.value = value

                # Aplicar formateo de moneda nativo de Excel si corresponde
                if currency_cols:
                    header_col_name = headers[col_idx - 1]
                    if header_col_name in currency_cols and cell.value is not None:
                        cell.number_format = '$#,##0.00'
            # Valores string
            else:
                cell.value = str(value)

            # Bordes para cuadrantes
            cell.border = cell_border

            # Zebra striping
            if apply_zebra:
                cell.fill = zebra_fill

            # Resaltar columna COD_CTA_SAFYC
            if safyc_col_idx == col_idx:
                cell.fill = safyc_fill

    # Ajustar ancho de columnas automáticamente
    for column_cells in ws.columns:
        length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    # Calcular longitud aproximada
                    cell_length = len(str(cell.value))
                    if cell_length > length:
                        length = cell_length
            except:
                pass

        # Establecer ancho (con un máximo y mínimo)
        adjusted_width = min(max(length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


def create_comparison_excels(
    df_completa: pd.DataFrame, df_existentes: pd.DataFrame
) -> tuple[bytes, bytes]:
    """
    Crea ambos archivos Excel de comparación.

    Args:
        df_completa: DataFrame con todas las diferencias (NUEVA, EXISTENTE, ELIMINADA)
        df_existentes: DataFrame con solo cuentas EXISTENTES

    Returns:
        tuple: (bytes_completa, bytes_existentes)
            - bytes_completa: Excel de comparación completa
            - bytes_existentes: Excel de solo existentes

    Example:
        >>> excel1, excel2 = create_comparison_excels(df_comp, df_exist)
        >>> with open('completa.xlsx', 'wb') as f:
        ...     f.write(excel1)
        >>> with open('existentes.xlsx', 'wb') as f:
        ...     f.write(excel2)
    """
    print(f"Generando Excel completo con {len(df_completa)} registros...")

    # Columnas a formatear como moneda en Excel
    currency_cols = {"INGRESOS", "EGRESOS", "SALDOFIN_2025", "SALDO_INI_2026", "DIFERENCIA"}

    # Trabajar sobre copias para no mutar los DataFrames originales
    df_comp_for_xlsx = df_completa.copy()
    df_exist_for_xlsx = df_existentes.copy()

    # Eliminar SALDO_INI sólo en las copias que se van a escribir
    df_comp_for_xlsx = df_comp_for_xlsx.drop(columns=["SALDO_INI"], errors="ignore")
    df_exist_for_xlsx = df_exist_for_xlsx.drop(columns=["SALDO_INI"], errors="ignore")

    excel_completa = dataframe_to_excel_bytes(df_comp_for_xlsx, "Sheet1", currency_cols=currency_cols)

    print(f"Generando Excel de existentes con {len(df_existentes)} registros...")
    excel_existentes = dataframe_to_excel_bytes(df_exist_for_xlsx, "Sheet1", currency_cols=currency_cols)

    return excel_completa, excel_existentes


def format_excel_with_styles(wb: Workbook) -> Workbook:
    """
    Aplica estilos adicionales a un workbook de Excel.
    (Función auxiliar para futuros mejoramientos)

    Args:
        wb: Workbook de openpyxl

    Returns:
        Workbook: Workbook con estilos aplicados
    """
    # Por ahora, solo retorna el workbook sin cambios
    # Aquí se pueden agregar estilos adicionales en el futuro:
    # - Colores alternados en filas
    # - Formatos de número personalizados
    # - Filtros automáticos
    # - Freeze panes
    return wb
